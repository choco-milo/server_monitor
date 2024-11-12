import pandas as pd 
import paramiko
from io import StringIO
import socket
import openpyxl 



def connect_to_server(hostname, username, password, commands):
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname=hostname, port=22, username=username,key_filename=password,  timeout=1)
        
        outputs = []
        for c in commands:
            stdin, stdout, stderr = client.exec_command(c)
            output = stdout.read().decode()
            outputs.append(output.strip())
        client.close()
        return outputs

    except paramiko.AuthenticationException:
        return "Unable to connect, please check private key"
    except (paramiko.SSHException, socket.timeout) as e:
        return f"Connection error: {str(e)}"
    except Exception as e:
        return f"{str(e)}"

def parse_df_output(output):
    try:
        df = pd.read_csv(StringIO(output), sep='\s+')
        return df
    except pd.errors.EmptyDataError:
        return None
    except Exception as e:
        print(f"Error parsing output: {e}")
        return None



def process_servers(servers_df, template_path, output_path):
    messages = [] 
    commands = [
        "df -h --total | grep 'total' | awk '{print $2}'",
        "df -h --total | grep 'total' | awk '{print $5}'" ,
        "lscpu | grep '^CPU(s)' | awk '{print $2}'",
        "top -bn1 | grep 'Cpu(s)' | sed -r 's/[%]+/ /g' | awk '{print 100-$8}'",
        "free -h | awk 'NR==2{print $2}'",
        "free | awk 'NR==2{print ($3/$2)*100}'"
    ]

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb['data']
        excel_row = 1

        for index, row in servers_df.iterrows():
            hostname = row['Servers']  
            username = row['Username']
            password = 'pk.txt'
        
            
            result = connect_to_server(hostname, username,password, commands)
            
            if isinstance(result, str):  
                messages.append(f"Failed to connect to {hostname, username}: {result}")
                continue  
            
            try:
                excel_row += 1
                disk_capacity,disk_usage, cpu_capa,cpu_usage,ram_capacity,ram_usage= result
                ws.cell(row=excel_row, column=1).value = hostname
                ws.cell(row=excel_row, column=2).value = disk_capacity
                ws.cell(row=excel_row, column=3).value = float(disk_usage.strip('%')) /100
                ws.cell(row=excel_row, column=4).value = float(cpu_capa)
                ws.cell(row=excel_row, column=5).value = float(cpu_usage) / 100                          
                ws.cell(row=excel_row, column=6).value = ram_capacity                           
                ws.cell(row=excel_row, column=7).value = float(ram_usage) / 100                   
                ws.cell(row=excel_row, column=3).number_format = '0%'  
                ws.cell(row=excel_row, column=5).number_format = '0.00%'  
                ws.cell(row=excel_row, column=7).number_format = '0.00%' 
            
            except Exception as e:
                    messages.append(f"Error processing data from {hostname}: {e}")
 
        wb.save(template_path) 
        wb.save(output_path)  
        return messages

    except Exception as e:
        return [f"Error updating Excel template: {str(e)}"]
